"""Crawler giá đối thủ - bản chuyển đổi từ crawl.gs (Google Apps Script) sang Python.

Toàn bộ logic tìm URL sản phẩm, trích tên/giá, upsert theo URL và xuất Excel
được giữ nguyên như bản gốc. Chỉ khác về nơi lưu trữ (một file Excel cục bộ
thay cho Google Sheets) và cách chạy (CLI + vòng lặp tự chờ thay cho trigger).
"""

import argparse
import gzip
import json
import math
import os
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import openpyxl
import requests
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

STATE_FILE = '.crawler_state.json'


# ---------------------------------------------------------------------------
# Cấu hình (tương đương CRAWLER_CONFIG trong crawl.gs)
# ---------------------------------------------------------------------------

def _sgt_fallback_listings():
    return [f'https://sgt.com.vn/collections/all?page={page}' for page in range(1, 51)]


def _dienmay88_fallback_listings():
    urls = ['https://dienmay88.vn/cua-hang/']
    urls += [f'https://dienmay88.vn/cua-hang/page/{page}/' for page in range(2, 101)]
    return urls


def _thienphu_fallback_listings():
    return [
        'https://dienmaythienphu.vn/tivi',
        'https://dienmaythienphu.vn/dieu-hoa',
        'https://dienmaythienphu.vn/tu-lanh',
        'https://dienmaythienphu.vn/may-giat',
        'https://dienmaythienphu.vn/dien-gia-dung',
        'https://dienmaythienphu.vn/tu-dong-tu-mat',
        'https://dienmaythienphu.vn/thiet-bi-nha-bep',
        'https://dienmaythienphu.vn/thiet-bi-am-thanh',
        'https://dienmaythienphu.vn/dieu-hoa-cong-trinh',
        'https://dienmaythienphu.vn/bom-nhiet-heat-pump',
    ]


def _dienmayabc_fallback_listings():
    categories = [
        'tivi.html', 'dieu-hoa.html', 'tu-lanh.html', 'may-giat.html',
        'may-say.html', 'may-loc-nuoc.html', 'may-rua-bat.html',
        'may-hut-am.html', 'gia-dung.html',
    ]
    urls = []
    for path in categories:
        urls.append(f'https://dienmayabc.com/{path}')
        urls += [f'https://dienmayabc.com/{path}?page={page}' for page in range(2, 31)]
    return urls


CONFIG = {
    # File Excel dùng làm "cơ sở dữ liệu": chứa CRAWL_QUEUE, CRAWL_LOG và
    # 4 sheet dữ liệu đối thủ - tương đương Google Sheet gốc.
    'workbook_path': 'crawl_data.xlsx',
    'queue_sheet': 'CRAWL_QUEUE',
    'log_sheet': 'CRAWL_LOG',

    'data_headers': ['STT', 'Tên sản phẩm', 'url', 'Giá sản phẩm', 'Giá sale'],

    # Xuất 4 sheet đối thủ thành một file Excel (.xlsx) riêng, lưu cục bộ.
    'excel_export': {
        'enabled': True,
        'output_dir': '.',
        'file_name': 'bang_gia_doi_thu.xlsx',
        # True: ghi đè file cũ (giữ nguyên tên). False: mỗi lần xuất tạo file có timestamp.
        'update_existing_file': True,
        # False: chỉ xuất khi crawl xong toàn bộ hàng đợi. True: xuất sau từng batch.
        'export_after_each_batch': False,
    },

    # Nên để thấp để không tạo quá nhiều request cùng lúc tới website.
    'batch_size': 10,
    'trigger_minutes': 5,

    # 0 = không giới hạn. Đặt 20 hoặc 100 để chạy thử trước.
    'max_product_urls_per_site': 0,

    'max_sitemap_depth': 4,
    'max_sitemaps_per_site': 100,

    'sites': [
        {
            'id': 'sgt',
            'name': 'SGT',
            'sheet_name': 'sgt.com.vn',
            'base_url': 'https://sgt.com.vn',
            'fallback_listings': _sgt_fallback_listings,
        },
        {
            'id': 'dienmay88',
            'name': 'Điện Máy 88',
            'sheet_name': 'dienmay88.vn',
            'base_url': 'https://dienmay88.vn',
            'fallback_listings': _dienmay88_fallback_listings,
        },
        {
            'id': 'thienphu',
            'name': 'Điện Máy Thiên Phú',
            'sheet_name': 'dienmaythienphu.vn',
            'base_url': 'https://dienmaythienphu.vn',
            'fallback_listings': _thienphu_fallback_listings,
        },
        {
            'id': 'dienmayabc',
            'name': 'Điện Máy ABC',
            'sheet_name': 'dienmayabc.com',
            'base_url': 'https://dienmayabc.com',
            'fallback_listings': _dienmayabc_fallback_listings,
        },
    ],
}


def get_site_by_id(site_id):
    for site in CONFIG['sites']:
        if site['id'] == site_id:
            return site
    return None


# ---------------------------------------------------------------------------
# Tiện ích chuỗi / HTML (tương đương phần cuối crawl.gs)
# ---------------------------------------------------------------------------

def decode_html(text):
    text = str(text)
    text = re.sub(r'&nbsp;|&#160;', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'&amp;', '&', text, flags=re.IGNORECASE)
    text = re.sub(r'&quot;', '"', text, flags=re.IGNORECASE)
    text = re.sub(r'&#39;|&apos;', "'", text, flags=re.IGNORECASE)
    text = re.sub(r'&lt;', '<', text, flags=re.IGNORECASE)
    text = re.sub(r'&gt;', '>', text, flags=re.IGNORECASE)
    text = re.sub(r'&#x2F;', '/', text, flags=re.IGNORECASE)
    text = re.sub(r'&#(\d+);', lambda m: chr(int(m.group(1))), text)
    return text


def decode_xml(text):
    return decode_html(text)


def strip_tags(html):
    text = decode_html(str(html))
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def html_to_text(html):
    text = decode_html(str(html))
    text = re.sub(r'<script\b[\s\S]*?</script>', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'<style\b[\s\S]*?</style>', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def clean_text(text):
    result = decode_html(str(text))
    result = re.sub(r'\s+', ' ', result)
    return result.strip()


def normalize_price(value):
    if value is None or value == '':
        return None

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return round(value) if math.isfinite(value) and value > 0 else None

    text = decode_html(str(value)).strip()
    if not text or re.search(r'liên\s*hệ', text, re.IGNORECASE):
        return None

    compact = re.sub(r'\s+', '', text)

    # JSON-LD đôi khi trả giá dạng 8390000.00. Không được xóa dấu chấm
    # theo cách biến nó thành 839000000.
    if re.match(r'^\d+[.,]\d{1,2}$', compact):
        decimal_price = float(compact.replace(',', '.'))
        return round(decimal_price) if math.isfinite(decimal_price) and decimal_price > 0 else None

    # Giá Việt Nam thường có dạng 8.390.000 hoặc 8,390,000.
    digits = re.sub(r'[^0-9]', '', compact)
    if not digits:
        return None

    price = int(digits)
    return price if price > 0 else None


def unique_(values):
    seen = set()
    result = []
    for value in values:
        key = str(value).strip()
        if key and key not in seen:
            seen.add(key)
            result.append(key)
    return result


def normalize_url(base_url, url):
    if not url:
        return ''
    normalized = str(url).strip()
    if re.match(r'^//', normalized):
        normalized = f'https:{normalized}'
    elif re.match(r'^/', normalized):
        normalized = f'{base_url}{normalized}'
    elif not re.match(r'^https?://', normalized, re.IGNORECASE):
        normalized = f"{base_url}/{normalized.lstrip('/')}"
    return re.sub(r'#.*$', '', normalized)


def get_url_path(url):
    match = re.match(r'^https?://[^/]+(/[^?#]*)?', str(url), re.IGNORECASE)
    if match and match.group(1):
        path = re.sub(r'/+$', '', match.group(1))
        return path or '/'
    return '/'


def is_excluded_path(path):
    return bool(re.search(
        r'/(?:gioi-thieu|lien-he|chinh-sach|tin-tuc|blog|blogs|tag|author|cart|'
        r'gio-hang|checkout|thanh-toan|tai-khoan|my-account|search|tim-kiem)(?:/|$)',
        path, re.IGNORECASE))


def is_known_abc_category(path):
    return bool(re.match(
        r'^/(?:tivi|dieu-hoa|tu-lanh|may-giat|may-say|may-loc-nuoc|may-rua-bat|'
        r'may-hut-am|gia-dung|lg|samsung|sony|toshiba|tcl|xiaomi|panasonic|sharp|'
        r'aqua|electrolux|funiki|bosch)\.html$',
        path, re.IGNORECASE))


# ---------------------------------------------------------------------------
# HTTP fetch (tương đương UrlFetchApp)
# ---------------------------------------------------------------------------

def build_headers():
    return {
        'User-Agent': 'Mozilla/5.0 (compatible; PythonPriceMonitor/1.0)',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'vi-VN,vi;q=0.9,en;q=0.7',
    }


def fetch_text_safe(url):
    try:
        response = requests.get(url, headers=build_headers(), allow_redirects=True, timeout=30)
    except requests.RequestException:
        return None

    if not (200 <= response.status_code < 300):
        return None

    if re.search(r'\.gz(?:\?|$)', url, re.IGNORECASE):
        try:
            return gzip.decompress(response.content).decode('utf-8', errors='replace')
        except OSError:
            return None

    return response.content.decode('utf-8', errors='replace')


def fetch_all(urls, max_workers=10):
    """Tương đương UrlFetchApp.fetchAll(): fetch song song, giữ nguyên thứ tự."""
    results = [None] * len(urls)

    def do_fetch(index, url):
        try:
            response = requests.get(url, headers=build_headers(), allow_redirects=True, timeout=30)
            return index, response.status_code, response.content, None
        except requests.RequestException as error:
            return index, None, None, error

    workers = max(1, min(max_workers, len(urls) or 1))
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = [executor.submit(do_fetch, i, u) for i, u in enumerate(urls)]
        for future in as_completed(futures):
            index, status_code, content, error = future.result()
            results[index] = (status_code, content, error)

    return results


# ---------------------------------------------------------------------------
# Tìm URL sản phẩm qua sitemap
# ---------------------------------------------------------------------------

def looks_like_product_sitemap(url):
    return bool(re.search(r'(?:product|products|san-pham|hang-hoa)', url, re.IGNORECASE))


def is_likely_product_url(site_id, url, source_sitemap):
    if not url or not re.match(r'^https?://', url, re.IGNORECASE):
        return False

    if re.search(r'\.(?:jpg|jpeg|png|gif|webp|pdf|xml|gz)(?:\?|$)', url, re.IGNORECASE):
        return False

    if looks_like_product_sitemap(source_sitemap):
        return True

    path = get_url_path(url)

    if site_id == 'sgt':
        return path.startswith('/products/')
    if site_id == 'dienmay88':
        return not is_excluded_path(path) and len([s for s in path.split('/') if s]) >= 1
    if site_id == 'thienphu':
        segments = [s for s in path.split('/') if s]
        return len(segments) >= 3 and not path.startswith('/tin-tuc/') and not is_excluded_path(path)
    if site_id == 'dienmayabc':
        return bool(re.search(r'\.html$', path, re.IGNORECASE)) and \
            not is_known_abc_category(path) and not is_excluded_path(path)
    return False


def extract_sitemap_locations(xml):
    locations = []
    for match in re.finditer(r'<loc\b[^>]*>([\s\S]*?)</loc>', xml, re.IGNORECASE):
        url = decode_xml(strip_tags(match.group(1))).strip()
        if url:
            locations.append(url)
    return unique_(locations)


def find_sitemap_urls(base_url):
    urls = []
    robots = fetch_text_safe(f'{base_url}/robots.txt')

    if robots:
        for match in re.finditer(r'^\s*Sitemap:\s*(\S+)\s*$', robots, re.IGNORECASE | re.MULTILINE):
            urls.append(decode_xml(match.group(1)))

    if not urls:
        candidates = [
            f'{base_url}/sitemap.xml',
            f'{base_url}/sitemap_index.xml',
            f'{base_url}/wp-sitemap.xml',
        ]
        for url in candidates:
            xml = fetch_text_safe(url)
            if xml and re.search(r'<(?:sitemapindex|urlset)\b', xml, re.IGNORECASE):
                urls.append(url)

    return unique_(urls)


def walk_sitemap(sitemap_url, site, depth, output, visited, counter):
    if depth > CONFIG['max_sitemap_depth']:
        return
    if sitemap_url in visited:
        return
    if counter['value'] >= CONFIG['max_sitemaps_per_site']:
        return

    visited[sitemap_url] = True
    counter['value'] += 1

    xml = fetch_text_safe(sitemap_url)
    if not xml:
        return

    locations = extract_sitemap_locations(xml)

    if re.search(r'<sitemapindex\b', xml, re.IGNORECASE):
        product_sitemaps = [loc for loc in locations if looks_like_product_sitemap(loc)]
        child_sitemaps = product_sitemaps if product_sitemaps else locations
        for child_url in child_sitemaps:
            walk_sitemap(child_url, site, depth + 1, output, visited, counter)
        return

    for url in locations:
        if is_likely_product_url(site['id'], url, sitemap_url):
            output.append(normalize_url(site['base_url'], url))


def discover_product_urls_from_sitemaps(site):
    sitemap_urls = find_sitemap_urls(site['base_url'])
    output = []
    visited = {}
    counter = {'value': 0}

    for sitemap_url in sitemap_urls:
        walk_sitemap(sitemap_url, site, 0, output, visited, counter)

    return unique_(output)


# ---------------------------------------------------------------------------
# Tìm URL sản phẩm qua trang danh mục (fallback listing)
# ---------------------------------------------------------------------------

def is_likely_product_url_from_listing(site_id, url, anchor_tag):
    path = get_url_path(url)

    if site_id == 'sgt':
        return path.startswith('/products/')
    if site_id == 'dienmay88':
        return bool(re.search(r'woocommerce-LoopProduct-link|product-item|product_type_', anchor_tag, re.IGNORECASE))
    if site_id == 'thienphu':
        segments = [s for s in path.split('/') if s]
        return len(segments) >= 3 and not path.startswith('/tin-tuc/')
    if site_id == 'dienmayabc':
        return bool(re.search(r'\.html$', path, re.IGNORECASE)) and not is_known_abc_category(path)
    return False


def extract_product_links_from_listing(site, html):
    links = []
    for match in re.finditer(r'<a\b[^>]*href=["\']([^"\']+)["\'][^>]*>', html, re.IGNORECASE):
        absolute_url = normalize_url(site['base_url'], decode_html(match.group(1)))
        if is_likely_product_url_from_listing(site['id'], absolute_url, match.group(0)):
            links.append(absolute_url)
    return unique_(links)


def discover_product_urls_from_listings(site):
    output = []
    listing_urls = site['fallback_listings']()
    empty_pages = 0

    for listing_url in listing_urls:
        html = fetch_text_safe(listing_url)

        if not html:
            empty_pages += 1
            if empty_pages >= 3:
                break
            continue

        links = extract_product_links_from_listing(site, html)

        if not links:
            empty_pages += 1
            if empty_pages >= 3 and site['id'] != 'thienphu':
                break
        else:
            empty_pages = 0
            output.extend(links)

        time.sleep(0.25)

    return unique_(output)


# ---------------------------------------------------------------------------
# Trích tên sản phẩm / giá từ trang chi tiết
# ---------------------------------------------------------------------------

def find_product_object(value):
    if value is None:
        return None

    if isinstance(value, list):
        for item in value:
            result = find_product_object(item)
            if result:
                return result
        return None

    if not isinstance(value, dict):
        return None

    type_value = value.get('@type')
    types = type_value if isinstance(type_value, list) else [type_value]
    if any(str(t).lower() == 'product' for t in types):
        return value

    for key in value:
        result = find_product_object(value[key])
        if result:
            return result

    return None


def extract_json_ld_product(html):
    pattern = r'<script\b[^>]*type=["\']application/ld\+json["\'][^>]*>([\s\S]*?)</script>'

    for match in re.finditer(pattern, html, re.IGNORECASE):
        json_text = match.group(1).strip()
        if not json_text:
            continue

        json_text = decode_html(json_text)
        json_text = re.sub(r'^\s*<!--', '', json_text)
        json_text = re.sub(r'-->\s*$', '', json_text)
        json_text = json_text.strip()

        try:
            data = json.loads(json_text)
        except ValueError:
            continue

        product = find_product_object(data)
        if product:
            return product

    return None


def get_offer_price(offers):
    if not offers:
        return None

    if isinstance(offers, list):
        for item in offers:
            price = get_offer_price(item)
            if price is not None:
                return price
        return None

    if not isinstance(offers, dict):
        return None

    for candidate in (offers.get('price'), offers.get('lowPrice'), offers.get('highPrice')):
        price = normalize_price(candidate)
        if price is not None:
            return price

    return None


def extract_first_price(text):
    match = re.search(r'\d{1,3}(?:[.,]\d{3}){1,3}|\d{5,12}', str(text))
    return normalize_price(match.group(0)) if match else None


def extract_price_by_regex(text, pattern, flags=re.IGNORECASE):
    match = re.search(pattern, str(text), flags)
    return normalize_price(match.group(1)) if match else None


def extract_price_from_tag(html, tag_name):
    pattern = rf'<{tag_name}\b[^>]*>([\s\S]*?)</{tag_name}>'
    match = re.search(pattern, str(html), re.IGNORECASE)
    return normalize_price(strip_tags(match.group(1))) if match else None


def extract_price_from_class(html, class_names):
    for class_name in class_names:
        escaped = re.escape(class_name)
        pattern = rf'<[^>]*class=["\'][^"\']*{escaped}[^"\']*["\'][^>]*>([\s\S]{{0,500}})'
        match = re.search(pattern, str(html), re.IGNORECASE)
        if match:
            price = extract_first_price(strip_tags(match.group(1)))
            if price is not None:
                return price
    return None


def extract_site_price_pair(html, site_id):
    text = html_to_text(html)

    if site_id == 'sgt':
        match = re.search(
            r'(\d{1,3}(?:[.,]\d{3})+)\s*₫\s*(\d{1,3}(?:[.,]\d{3})+)\s*₫\s*-\d+%', text, re.IGNORECASE)
        if match:
            return normalize_price(match.group(2)), normalize_price(match.group(1))

    elif site_id == 'thienphu':
        match = re.search(
            r'(\d{1,3}(?:\.\d{3})+)\s*(?:₫)?\s+(\d{1,3}(?:\.\d{3})+)\s*(?:₫)?\s*-\d+%', text, re.IGNORECASE)
        if match:
            return normalize_price(match.group(2)), normalize_price(match.group(1))

    elif site_id == 'dienmayabc':
        current = extract_price_by_regex(text, r'Giá\s*ABC\s*:\s*([\d.,]+)')
        original = extract_price_by_regex(
            text, r'(?:Giá\s*(?:thị trường|niêm yết|gốc)|Giá cũ)\s*:\s*([\d.,]+)')
        regular = original if original is not None else current
        sale = current if (original is not None and current is not None and original > current) else None
        return regular, sale

    return None, None


def extract_meta_content(html, attribute_name, attribute_value):
    def extract_attribute(tag, name):
        escaped = re.escape(name)
        match = re.search(rf'\b{escaped}\s*=\s*(["\'])([\s\S]*?)\1', tag, re.IGNORECASE)
        return decode_html(match.group(2)) if match else ''

    for tag in re.findall(r'<meta\b[^>]*>', html, re.IGNORECASE):
        actual_value = extract_attribute(tag, attribute_name)
        if actual_value and actual_value.lower() == attribute_value.lower():
            return extract_attribute(tag, 'content') or ''
    return ''


def extract_meta_price(html):
    keys = [
        ('property', 'product:price:amount'),
        ('itemprop', 'price'),
        ('name', 'product:price:amount'),
    ]
    for attr_name, attr_value in keys:
        price = normalize_price(extract_meta_content(html, attr_name, attr_value))
        if price is not None:
            return price
    return None


def extract_product_name(html):
    match = re.search(r'<h1\b[^>]*>([\s\S]*?)</h1>', html, re.IGNORECASE)
    if match:
        return strip_tags(match.group(1))

    return (extract_meta_content(html, 'property', 'og:title')
            or extract_meta_content(html, 'name', 'twitter:title')
            or '')


def extract_site_price(html, site_id):
    if site_id == 'dienmayabc':
        return extract_price_by_regex(
            html, r'Giá\s*ABC\s*:\s*(?:<[^>]+>\s*)*([\d.,]+)\s*(?:₫|đ|VND)')

    if site_id == 'thienphu':
        visible_text = html_to_text(html)
        return extract_price_by_regex(visible_text, r'(?:^|\s)Giá\s*:\s*([\d.]{5,})')

    if site_id == 'dienmay88':
        amount_matches = re.findall(
            r'<[^>]*class=["\'][^"\']*woocommerce-Price-amount[^"\']*["\'][^>]*>[\s\S]*?</[^>]+>',
            html, re.IGNORECASE)
        for match in amount_matches:
            price = normalize_price(strip_tags(match))
            if price is not None:
                return price
        return None

    if site_id == 'sgt':
        visible_text = html_to_text(html)
        patterns = [
            r'Giá\s*Sốc\s*Online\s*([\d.,]+)\s*₫',
            r'([\d.,]+)\s*₫',
        ]
        for pattern in patterns:
            price = extract_price_by_regex(visible_text, pattern)
            if price is not None:
                return price
        return None

    return None


def extract_product_price_pair(html, site_id, json_ld_product):
    regular_price = extract_price_from_tag(html, 'del')
    sale_price = extract_price_from_tag(html, 'ins')

    if regular_price is None:
        regular_price = extract_price_from_class(html, [
            'old-price', 'price-old', 'old_price', 'compare-price',
            'compare_at_price', 'list-price', 'regular-price',
            'price-regular', 'market-price', 'price_market',
        ])

    if sale_price is None:
        sale_price = extract_price_from_class(html, [
            'sale-price', 'price-sale', 'special-price', 'current-price',
            'price-current', 'new-price', 'price-new', 'product-price-sale',
        ])

    site_regular, site_sale = extract_site_price_pair(html, site_id)

    if regular_price is None:
        regular_price = site_regular
    if sale_price is None:
        sale_price = site_sale

    current_price = get_offer_price(json_ld_product.get('offers') if json_ld_product else None)

    if current_price is None:
        current_price = extract_meta_price(html)

    if current_price is None:
        current_price = extract_site_price(html, site_id)

    # Nếu có giá gốc và giá hiện tại thì giá nhỏ hơn là giá sale.
    if regular_price is not None and sale_price is not None:
        if regular_price < sale_price:
            regular_price, sale_price = sale_price, regular_price
        elif regular_price == sale_price:
            sale_price = None
    elif regular_price is not None and current_price is not None:
        if regular_price > current_price:
            sale_price = current_price
        elif current_price > regular_price:
            sale_price = regular_price
            regular_price = current_price
    elif sale_price is not None and current_price is not None:
        if sale_price > current_price:
            regular_price = sale_price
            sale_price = current_price
        elif sale_price < current_price:
            regular_price = current_price
        else:
            regular_price = current_price
            sale_price = None
    elif regular_price is None and sale_price is None:
        regular_price = current_price
    elif regular_price is None and sale_price is not None:
        regular_price = sale_price
        sale_price = None

    return regular_price, sale_price


def is_confirmed_product_page(html, site_id, url):
    path = get_url_path(url)

    if site_id == 'sgt':
        return path.startswith('/products/')
    if site_id == 'dienmay88':
        return bool(re.search(r'single-product|product_title|woocommerce-product-details', html, re.IGNORECASE))
    if site_id == 'thienphu':
        segments = [s for s in path.split('/') if s]
        return (len(segments) >= 3 and not path.startswith('/tin-tuc/')
                and bool(re.search(r'(?:Giá\s*:|Gọi đặt mua|Tình trạng|Tình trang)',
                                    html_to_text(html), re.IGNORECASE)))
    if site_id == 'dienmayabc':
        return bool(re.search(r'Giá\s*ABC\s*:', html_to_text(html), re.IGNORECASE))
    return False


def page_has_contact_price(html):
    text = html_to_text(html)
    # Chỉ coi là giá liên hệ khi từ "Liên hệ" nằm ngay sau nhãn giá,
    # tránh bắt nhầm chữ "Liên hệ" ở menu hoặc chân trang.
    return bool(re.search(r'(?:Giá(?:\s*ABC)?|Giá bán|Giá sản phẩm)\s*:?\s*Liên hệ', text, re.IGNORECASE))


def build_product_result(website, site_id, url, name, regular_price, sale_price, contact_price):
    return {
        'website': website,
        'name': clean_text(name),
        'url': url,
        'regular_price': regular_price,
        'sale_price': sale_price,
        'contact_price': contact_price is True,
        'crawled_at': datetime.now(),
        'site_id': site_id,
    }


def extract_product(html, site_id, website, url):
    json_ld_product = extract_json_ld_product(html)
    name = (json_ld_product.get('name') if json_ld_product and json_ld_product.get('name')
            else extract_product_name(html))

    if not name or not is_confirmed_product_page(html, site_id, url):
        return None

    regular_price, sale_price = extract_product_price_pair(html, site_id, json_ld_product)
    contact_price = page_has_contact_price(html)

    if regular_price is None and sale_price is None and not contact_price:
        return None

    return build_product_result(website, site_id, url, name, regular_price, sale_price, contact_price)


# ---------------------------------------------------------------------------
# Excel workbook helpers (thay cho Google Sheets)
# ---------------------------------------------------------------------------

def load_or_create_workbook(path):
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    return openpyxl.Workbook()


def save_workbook(wb):
    wb.save(CONFIG['workbook_path'])


def get_or_create_sheet(wb, name):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)


def remove_default_sheet(wb):
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb['Sheet']


def clear_sheet_rows(ws):
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)


def set_header(ws, headers):
    for col_index, title in enumerate(headers, start=1):
        ws.cell(row=1, column=col_index, value=title)


def apply_header_and_widths(ws, widths):
    ws.freeze_panes = 'A2'
    for col_index in range(1, len(widths) + 1):
        ws.cell(row=1, column=col_index).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col_index)].width = widths[col_index - 1]


def format_data_sheet(ws):
    apply_header_and_widths(ws, [8, 45, 55, 16, 16])
    if ws.max_row >= 2:
        for row in range(2, ws.max_row + 1):
            for col in (4, 5):
                ws.cell(row=row, column=col).number_format = '#,##0'


def ensure_data_sheet(wb, sheet_name):
    ws = get_or_create_sheet(wb, sheet_name)
    set_header(ws, CONFIG['data_headers'])
    format_data_sheet(ws)
    return ws


def renumber_sheet(ws):
    if ws.max_row < 2:
        return
    for offset, row in enumerate(range(2, ws.max_row + 1), start=1):
        ws.cell(row=row, column=1, value=offset)


def log_(wb, level, website, message):
    ws = get_or_create_sheet(wb, CONFIG['log_sheet'])
    ws.append([datetime.now(), level, website, message])


def mark_queue_row(ws, row, status, error):
    ws.cell(row=row, column=4, value=status)
    ws.cell(row=row, column=5, value=error or '')
    ws.cell(row=row, column=6, value=datetime.now())


def count_pending_rows(ws):
    if ws.max_row < 2:
        return 0
    count = 0
    for row in range(2, ws.max_row + 1):
        status = ws.cell(row=row, column=4).value
        if str(status or '').upper() == 'PENDING':
            count += 1
    return count


# ---------------------------------------------------------------------------
# Trạng thái nhỏ giữa các lần chạy (thay cho PropertiesService)
# ---------------------------------------------------------------------------

def load_state():
    if not os.path.exists(STATE_FILE):
        return {}
    with open(STATE_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)


def save_state(data):
    state = load_state()
    state.update(data)
    with open(STATE_FILE, 'w', encoding='utf-8') as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


def show_last_export():
    state = load_state()
    path = state.get('last_excel_file_path')
    if not path:
        raise RuntimeError('Chưa có file Excel nào được xuất.')
    print(path)
    return path


# ---------------------------------------------------------------------------
# BƯỚC 1: Khởi tạo hàng đợi
# ---------------------------------------------------------------------------

def setup_price_crawler(wb):
    queue_ws = get_or_create_sheet(wb, CONFIG['queue_sheet'])
    log_ws = get_or_create_sheet(wb, CONFIG['log_sheet'])
    remove_default_sheet(wb)

    clear_sheet_rows(queue_ws)
    clear_sheet_rows(log_ws)

    set_header(queue_ws, ['Site ID', 'Website', 'URL', 'Status', 'Lỗi', 'Cập nhật lúc'])
    set_header(log_ws, ['Thời gian', 'Mức', 'Website', 'Nội dung'])

    # Không xóa dữ liệu cũ. Nếu sheet đã tồn tại thì dữ liệu sẽ được update theo URL.
    for site in CONFIG['sites']:
        ensure_data_sheet(wb, site['sheet_name'])

    queue_rows = []

    for site in CONFIG['sites']:
        log_(wb, 'INFO', site['name'], 'Bắt đầu tìm URL sản phẩm.')

        urls = discover_product_urls_from_sitemaps(site)

        if not urls:
            log_(wb, 'WARN', site['name'], 'Không lấy được URL từ sitemap, chuyển sang trang danh mục.')
            urls = discover_product_urls_from_listings(site)

        urls = unique_(urls)

        if CONFIG['max_product_urls_per_site'] > 0:
            urls = urls[:CONFIG['max_product_urls_per_site']]

        for url in urls:
            queue_rows.append([site['id'], site['name'], url, 'PENDING', '', datetime.now()])

        log_(wb, 'INFO', site['name'], f"Đã đưa {len(urls)} URL vào hàng đợi.")

    for offset, row in enumerate(queue_rows, start=2):
        for col_index, value in enumerate(row, start=1):
            queue_ws.cell(row=offset, column=col_index, value=value)

    apply_header_and_widths(queue_ws, [10, 22, 60, 12, 30, 20])
    apply_header_and_widths(log_ws, [20, 8, 22, 60])

    save_workbook(wb)

    print(f"Đã tạo hàng đợi {len(queue_rows)} URL. Dữ liệu cũ được giữ lại và sẽ update theo URL.")


# ---------------------------------------------------------------------------
# BƯỚC 2: Crawl một batch URL đang PENDING
# ---------------------------------------------------------------------------

def crawl_next_batch(wb):
    if CONFIG['queue_sheet'] not in wb.sheetnames:
        raise RuntimeError('Chưa có hàng đợi. Hãy chạy setup_price_crawler() trước.')

    queue_ws = wb[CONFIG['queue_sheet']]
    if queue_ws.max_row < 2:
        raise RuntimeError('Chưa có hàng đợi. Hãy chạy setup_price_crawler() trước.')

    pending = []
    for row in range(2, queue_ws.max_row + 1):
        status = queue_ws.cell(row=row, column=4).value
        if str(status or '').upper() == 'PENDING':
            pending.append({
                'row': row,
                'site_id': queue_ws.cell(row=row, column=1).value,
                'website': queue_ws.cell(row=row, column=2).value,
                'url': queue_ws.cell(row=row, column=3).value,
            })
            if len(pending) >= CONFIG['batch_size']:
                break

    if not pending:
        if CONFIG['excel_export']['enabled']:
            export_products_to_excel(wb)
        print('Đã crawl xong và xuất file Excel.')
        return 0

    urls = [item['url'] for item in pending]

    try:
        responses = fetch_all(urls, max_workers=CONFIG['batch_size'])
    except Exception as error:
        for item in pending:
            mark_queue_row(queue_ws, item['row'], 'ERROR', str(error))
        save_workbook(wb)
        raise

    products = []

    for item, (status_code, content, error) in zip(pending, responses):
        if error is not None:
            mark_queue_row(queue_ws, item['row'], 'ERROR', str(error))
            continue

        if status_code is None or not (200 <= status_code < 300):
            mark_queue_row(queue_ws, item['row'], 'ERROR', f'HTTP {status_code}')
            continue

        try:
            html = content.decode('utf-8', errors='replace')
            product = extract_product(html, item['site_id'], item['website'], item['url'])

            if not product:
                mark_queue_row(
                    queue_ws, item['row'], 'SKIP',
                    'Không nhận diện được trang sản phẩm hoặc không tìm thấy tên/giá.')
                continue

            products.append(product)
            mark_queue_row(queue_ws, item['row'], 'DONE', '')
        except Exception as error:
            mark_queue_row(queue_ws, item['row'], 'ERROR', str(error))

    if products:
        upsert_products(wb, products)

    remaining = count_pending_rows(queue_ws)
    print(f"Đã xử lý {len(pending)} URL, còn {remaining} URL.")

    if remaining == 0:
        if CONFIG['excel_export']['enabled']:
            export_products_to_excel(wb)
    elif CONFIG['excel_export']['enabled'] and CONFIG['excel_export']['export_after_each_batch']:
        export_products_to_excel(wb)

    save_workbook(wb)
    return remaining


def upsert_products(wb, products):
    products_by_site = {}
    for product in products:
        products_by_site.setdefault(product['site_id'], []).append(product)

    for site_id, site_products in products_by_site.items():
        site = get_site_by_id(site_id)
        if not site:
            continue

        ws = ensure_data_sheet(wb, site['sheet_name'])

        url_to_row = {}
        if ws.max_row >= 2:
            for row in range(2, ws.max_row + 1):
                url_value = ws.cell(row=row, column=3).value
                if url_value:
                    url_to_row[str(url_value).strip()] = row

        append_rows = []

        for product in site_products:
            if product['contact_price'] and product['regular_price'] is None:
                regular_value = 'Liên hệ'
            else:
                regular_value = '' if product['regular_price'] is None else product['regular_price']
            sale_value = '' if product['sale_price'] is None else product['sale_price']

            existing_row = url_to_row.get(product['url'])

            if existing_row:
                current_stt = ws.cell(row=existing_row, column=1).value
                ws.cell(row=existing_row, column=1, value=current_stt)
                ws.cell(row=existing_row, column=2, value=product['name'])
                ws.cell(row=existing_row, column=3, value=product['url'])
                ws.cell(row=existing_row, column=4, value=regular_value)
                ws.cell(row=existing_row, column=5, value=sale_value)
            else:
                append_rows.append([0, product['name'], product['url'], regular_value, sale_value])

        start_row = ws.max_row + 1
        for offset, row in enumerate(append_rows):
            for col_index, value in enumerate(row, start=1):
                ws.cell(row=start_row + offset, column=col_index, value=value)

        renumber_sheet(ws)
        format_data_sheet(ws)


# ---------------------------------------------------------------------------
# Xuất Excel tổng hợp 4 sheet đối thủ
# ---------------------------------------------------------------------------

def build_export_path(config):
    file_name = str(config.get('file_name') or 'bang_gia_doi_thu.xlsx').strip()
    if not file_name.lower().endswith('.xlsx'):
        file_name += '.xlsx'

    output_dir = config.get('output_dir') or '.'
    os.makedirs(output_dir, exist_ok=True)

    if config.get('update_existing_file'):
        return os.path.join(output_dir, file_name)

    base_name = file_name[:-len('.xlsx')]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return os.path.join(output_dir, f'{base_name}_{timestamp}.xlsx')


def export_products_to_excel(wb):
    config = CONFIG['excel_export']
    if not config['enabled']:
        return None

    for site in CONFIG['sites']:
        ensure_data_sheet(wb, site['sheet_name'])

    export_wb = openpyxl.Workbook()

    for site in CONFIG['sites']:
        source_ws = wb[site['sheet_name']]
        target_ws = export_wb.create_sheet(site['sheet_name'])
        for row in source_ws.iter_rows():
            for cell in row:
                target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
        format_data_sheet(target_ws)

    remove_default_sheet(export_wb)

    output_path = build_export_path(config)
    export_wb.save(output_path)

    save_state({
        'last_excel_file_path': str(output_path),
        'last_excel_export_at': datetime.now().isoformat(),
    })

    log_(wb, 'INFO', 'Local Export', f'Đã cập nhật Excel: {output_path}')
    print(f'Đã cập nhật Excel: {output_path}')

    return str(output_path)


# ---------------------------------------------------------------------------
# Reset / chạy toàn bộ quy trình
# ---------------------------------------------------------------------------

def reset_price_crawler(wb):
    for name in (CONFIG['queue_sheet'], CONFIG['log_sheet']):
        if name in wb.sheetnames:
            clear_sheet_rows(wb[name])
    save_workbook(wb)


def run_full_price_crawler(wb):
    setup_price_crawler(wb)
    remaining = crawl_next_batch(wb)

    while remaining:
        wait_minutes = CONFIG['trigger_minutes']
        print(f'Chờ {wait_minutes} phút trước khi crawl batch tiếp theo...')
        time.sleep(wait_minutes * 60)
        remaining = crawl_next_batch(wb)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description='Crawler giá đối thủ (bản Python, chuyển đổi từ crawl.gs)')
    parser.add_argument(
        'command',
        choices=['setup', 'batch', 'run', 'export', 'reset', 'last-export'],
        help=(
            'setup: khởi tạo hàng đợi | batch: crawl 1 batch | '
            'run: chạy toàn bộ (setup + lặp batch tới khi xong) | '
            'export: xuất Excel ngay | reset: xóa hàng đợi/log | '
            'last-export: in đường dẫn file Excel xuất gần nhất'
        ),
    )
    args = parser.parse_args()

    if args.command == 'last-export':
        show_last_export()
        return

    wb = load_or_create_workbook(CONFIG['workbook_path'])

    if args.command == 'setup':
        setup_price_crawler(wb)
    elif args.command == 'batch':
        crawl_next_batch(wb)
    elif args.command == 'run':
        run_full_price_crawler(wb)
    elif args.command == 'export':
        export_products_to_excel(wb)
        save_workbook(wb)
    elif args.command == 'reset':
        reset_price_crawler(wb)


if __name__ == '__main__':
    main()
